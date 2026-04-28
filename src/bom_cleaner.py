import os
import re
from dataclasses import dataclass
from pathlib import Path
import math

import pandas as pd
from openpyxl import load_workbook


STANDARD_COLUMNS = [
    "source_file",
    "source_row",
    "designators",
    "qty_per_board",
    "build_quantity",
    "overage_pct",
    "required_qty",
    "manufacturer",
    "mpn",
    "description",
    "supplier",
    "supplier_part_number",
    "unit_price",
    "subtotal",
    "lifecycle_status",
    "build_multiplier",
    "status",
    "notes",
]

MAPPABLE_COLUMNS = [
    "designators",
    "qty_per_board",
    "manufacturer",
    "mpn",
    "description",
    "supplier",
    "supplier_part_number",
    "unit_price",
    "subtotal",
    "lifecycle_status",
]

REVIEW_STATUSES = {
    "missing_mpn",
    "missing_manufacturer",
    "qty_mismatch",
    "manual_review",
}

WARNING_COLUMNS = ["source_file", "source_row", "field", "message", "value"]

COLUMN_ALIASES = {
    "designators": [
        "designator",
        "designators",
        "refdes",
        "ref_des",
        "reference",
        "references",
        "reference_designator",
        "reference_designators",
        "refs",
    ],
    "qty_per_board": [
        "qty",
        "quantity",
        "q'ty",
        "bom_qty",
        "quantity_per_board",
        "qty_per_board",
        "quantity_per_brd",
        "qty_per_brd",
    ],
    "manufacturer": [
        "manufacturer",
        "manufacturer_1",
        "mfg",
        "mfg_1",
        "mfr",
        "mfr_1",
        "maker",
        "brand",
        "manufacturer_name",
    ],
    "mpn": [
        "manufacturer_part_number",
        "manufacturer_part_number_1",
        "mfg_pn",
        "mfg_pn_1",
        "mfr_pn",
        "mfr_pn_1",
        "manufacturer_pn",
        "manufacturer_pn_1",
        "mpn",
        "part_number",
        "part_number_1",
        "part_no",
        "part_no_1",
        "pn",
    ],
    "description": [
        "description",
        "item_description",
        "desc",
        "value",
        "comment",
        "comments",
        "notes",
        "name",
    ],
    "supplier": [
        "supplier",
        "supplier_1",
        "vendor",
        "vendor_1",
        "distributor",
        "distributor_1",
    ],
    "supplier_part_number": [
        "supplier_part_number",
        "supplier_part_number_1",
        "supplier_pn",
        "supplier_pn_1",
        "supplier_part_no",
        "supplier_part_no_1",
        "vendor_part_number",
        "vendor_part_number_1",
        "digikey_part_number",
        "mouser_part_number",
    ],
    "unit_price": [
        "supplier_unit_price",
        "supplier_unit_price_1",
        "unit_price",
        "unit_cost",
        "price",
        "cost",
    ],
    "subtotal": [
        "supplier_subtotal",
        "supplier_subtotal_1",
        "subtotal",
        "extended_price",
        "extended_cost",
        "line_total",
        "total",
    ],
    "lifecycle_status": [
        "manufacturer_lifecycle",
        "manufacturer_lifecycle_1",
        "lifecycle",
        "lifecycle_status",
        "part_lifecycle",
    ],
}

ENCODING_REPLACEMENTS = {
    "\u00c2\u00b1": "\u00b1",
    "\u00c2\u00b0": "\u00b0",
    "\u00c2\u00b5": "\u00b5",
    "\u00ce\u00a9": "\u03a9",
    "\u00e2\u0084\u00a6": "\u03a9",
    "\u00c3\u0097": "x",
    "\u00ef\u00bf\u00bd": "\u00b1",
    "\ufffd": "\u00b1",
    "\u00c2": "",
}

DESIGNATOR_RANGE_RE = re.compile(r"^([A-Za-z]+)(\d+)\s*[-:]\s*\1?(\d+)$")
MULTIPLIER_COLUMN_RE = re.compile(r"^\d+x$")
GENERIC_MANUFACTURER_PREFIXES = {
    "cap",
    "capacitor",
    "conn",
    "connector",
    "fuse",
    "heat sink",
    "ind",
    "inductor",
    "res",
    "resistor",
    "test point",
}


@dataclass
class BomCleanResult:
    clean_bom: pd.DataFrame
    review_items: pd.DataFrame
    column_mapping: pd.DataFrame
    warnings: pd.DataFrame
    mapped_columns: dict


def load_bom(file_path):
    """
    Load a BOM file from CSV or Excel and return a pandas DataFrame.
    The cleaner treats incoming values as text so formatting survives cleanup.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    extension = Path(file_path).suffix.lower()

    if extension == ".csv":
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
            try:
                return pd.read_csv(file_path, dtype=str, keep_default_na=False, encoding=encoding)
            except UnicodeDecodeError:
                continue
        return pd.read_csv(file_path, dtype=str, keep_default_na=False)

    if extension == ".xlsx":
        return pd.read_excel(file_path, dtype=str, keep_default_na=False)

    raise ValueError("Unsupported file type. Use CSV or XLSX.")


def normalize_columns(df):
    """
    Normalize source headers while preserving the original header text in attrs.
    Duplicate-looking columns are made unique instead of being overwritten.
    """
    df = df.copy()
    original_columns = ["" if pd.isna(col) else str(col) for col in df.columns]
    normalized_columns = [_normalize_header(col) for col in original_columns]
    unique_columns = _make_unique_columns(normalized_columns)

    df.columns = unique_columns
    df.attrs["original_columns"] = dict(zip(unique_columns, original_columns))
    return df


def map_columns(df):
    """
    Backward-compatible helper for the older CLI flow.
    New code should prefer process_bom_file(), which also returns warnings
    and the Excel-ready review/mapping tables.
    """
    working_df, _ = _drop_junk_columns(df, source_file="")
    candidates = _find_column_candidates(working_df.columns)
    clean_df, _ = _build_standard_rows(working_df, candidates, source_file="")
    mapped = _build_mapped_columns(working_df, candidates)
    return clean_df, mapped


def clean_text(value):
    if pd.isna(value):
        return ""

    text = str(value)
    text = _clean_encoding_noise(text)
    text = text.replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()

    if text.lower() in {"nan", "none", "null"}:
        return ""

    return text.strip(" ,")


def count_designators(designator_text):
    """
    Count comma, semicolon, or space separated reference designators.
    Simple ranges like R1-R3 are counted as three designators.
    """
    text = clean_text(designator_text)
    if not text:
        return 0

    text = text.replace(";", ",")
    tokens = []
    for comma_part in text.split(","):
        tokens.extend(part for part in re.split(r"\s+", comma_part.strip()) if part)

    count = 0
    for token in tokens:
        range_match = DESIGNATOR_RANGE_RE.match(token)
        if range_match:
            start = int(range_match.group(2))
            end = int(range_match.group(3))
            if end >= start:
                count += end - start + 1
                continue
        count += 1

    return count


def infer_quantity(row):
    qty = clean_text(row.get("qty_per_board", ""))
    designators = clean_text(row.get("designators", ""))
    parsed_qty = _parse_quantity(qty)

    if parsed_qty is not None:
        return parsed_qty

    inferred_count = count_designators(designators)
    if inferred_count > 0:
        return inferred_count

    return qty


def determine_status(row):
    mpn = clean_text(row.get("mpn", ""))
    manufacturer = clean_text(row.get("manufacturer", ""))
    notes = clean_text(row.get("notes", "")).lower()

    if not mpn:
        return "missing_mpn"
    if "qty mismatch" in notes:
        return "qty_mismatch"
    if not manufacturer:
        return "missing_manufacturer"
    if "manual review" in notes:
        return "manual_review"
    if "combined field parsed" in notes:
        return "combined_field_parsed"
    return "clean"


def clean_rows(df):
    """
    Backward-compatible row cleaner. The full pipeline uses the private
    variant so it can return row-level warnings as a separate sheet.
    """
    clean_df, _ = _clean_standard_rows(df, source_file="")
    return clean_df


def preview_bom(df, num_rows=10):
    print("\nCleaned BOM Columns:")
    print(list(df.columns))

    print(f"\nTotal Rows: {len(df)}")

    if "status" in df.columns:
        print("\nStatus Counts:")
        print(df["status"].value_counts(dropna=False).to_string())

    print(f"\nPreview (first {num_rows} rows):")
    print(df.head(num_rows))


def process_bom_file(file_path):
    """
    Full Phase 1 workflow:
    load -> normalize headers -> ignore junk -> map fields -> validate rows.
    """
    source_file = os.path.basename(file_path)
    raw_df = load_bom(file_path)
    normalized_df = normalize_columns(raw_df)
    normalized_df["__source_row__"] = range(2, len(normalized_df) + 2)

    working_df, junk_warnings = _drop_junk_columns(normalized_df, source_file=source_file)
    working_df, empty_row_warnings = _drop_empty_rows(working_df, source_file=source_file)

    candidates = _find_column_candidates(working_df.columns)
    mapping_df = _build_column_mapping_df(working_df, candidates, source_file)
    mapped_columns = _build_mapped_columns(working_df, candidates)

    clean_df, mapping_warnings = _build_standard_rows(working_df, candidates, source_file)
    clean_df, row_warnings = _clean_standard_rows(clean_df, source_file)
    review_items = clean_df[clean_df["status"].isin(REVIEW_STATUSES)].copy()

    warnings_df = _make_warnings_df(
        junk_warnings + empty_row_warnings + mapping_warnings + row_warnings
    )

    return BomCleanResult(
        clean_bom=clean_df,
        review_items=review_items,
        column_mapping=mapping_df,
        warnings=warnings_df,
        mapped_columns=mapped_columns,
    )

def apply_project_quantities(df, build_quantity, overage_pct):
    """
    Add project-level quantity math to the cleaned BOM.
    required_qty = ceil(qty_per_board * build_quantity * (1 + overage_pct / 100))
    """
    updated_df = df.copy().astype(object)

    updated_df["build_quantity"] = pd.Series([None] * len(updated_df), dtype=object)
    updated_df["overage_pct"] = pd.Series([None] * len(updated_df), dtype=object)
    updated_df["required_qty"] = pd.Series([None] * len(updated_df), dtype=object)


    for index, row in updated_df.iterrows():
        qty_value = row.get("qty_per_board", "")
        parsed_qty = _parse_quantity(qty_value)

        updated_df.at[index, "build_quantity"] = build_quantity
        updated_df.at[index, "overage_pct"] = overage_pct

        if parsed_qty is None:
            notes = _split_notes(row.get("notes", ""))
            notes.append("manual review: required quantity not calculated")
            updated_df.at[index, "notes"] = "; ".join(_dedupe_notes(notes))

            if updated_df.at[index, "status"] == "clean":
                updated_df.at[index, "status"] = "manual_review"

            updated_df.at[index, "required_qty"] = ""
            continue

        required_qty = math.ceil(parsed_qty * build_quantity * (1 + overage_pct / 100))
        updated_df.at[index, "required_qty"] = required_qty

    return updated_df

def build_output_paths(
    file_path,
    output_folder,
    workbook_suffix="_cleaned_bom.xlsx",
    csv_suffix="_cleaned_bom.csv",
):
    source_stem = Path(file_path).stem
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", source_stem).strip("_") or "cleaned_bom"
    workbook_path = os.path.join(output_folder, f"{safe_stem}{workbook_suffix}")
    csv_path = os.path.join(output_folder, f"{safe_stem}{csv_suffix}")
    return workbook_path, csv_path


def export_clean_bom(df, output_path):
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    df.to_csv(output_path, index=False)


def export_clean_bom_workbook(result, output_path):
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        result.clean_bom.to_excel(writer, sheet_name="Clean_BOM", index=False)
        result.review_items.to_excel(writer, sheet_name="Review_Items", index=False)
        result.column_mapping.to_excel(writer, sheet_name="Column_Mapping", index=False)
        result.warnings.to_excel(writer, sheet_name="Warnings", index=False)

    _format_workbook(output_path)


def _normalize_header(header):
    text = clean_text(header).lower()
    text = text.replace("#", " number ")
    text = re.sub(r"[\r\n]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def _make_unique_columns(columns):
    counts = {}
    unique_columns = []

    for col in columns:
        base = col or "blank_column"
        if base not in counts:
            counts[base] = 0
            unique_columns.append(base)
            continue

        counts[base] += 1
        unique_columns.append(f"{base}_{counts[base]}")

    return unique_columns


def _clean_encoding_noise(text):
    cleaned = text

    if any(marker in cleaned for marker in ("\u00c2", "\u00c3", "\u00e2", "\u00ef")):
        try:
            cleaned = cleaned.encode("latin1").decode("utf-8")
        except UnicodeError:
            pass

    for bad, good in ENCODING_REPLACEMENTS.items():
        cleaned = cleaned.replace(bad, good)

    return cleaned


def _drop_junk_columns(df, source_file):
    keep_columns = []
    warnings = []
    original_columns = df.attrs.get("original_columns", {})

    for col in df.columns:
        if col == "__source_row__":
            keep_columns.append(col)
            continue

        original = original_columns.get(col, col)
        series = df[col].map(clean_text)

        if col.startswith("unnamed") or col.startswith("blank_column"):
            warnings.append(_warning(source_file, "", col, "Ignored junk/blank column", original))
            continue

        if series.eq("").all():
            warnings.append(_warning(source_file, "", col, "Ignored empty column", original))
            continue

        keep_columns.append(col)

    cleaned_df = df[keep_columns].copy()
    cleaned_df.attrs["original_columns"] = {
        col: original_columns.get(col, col) for col in keep_columns
    }
    return cleaned_df, warnings


def _drop_empty_rows(df, source_file):
    data_columns = [col for col in df.columns if col != "__source_row__"]
    keep_indexes = []
    warnings = []

    for index, row in df.iterrows():
        if any(clean_text(row.get(col, "")) for col in data_columns):
            keep_indexes.append(index)
            continue

        source_row = row.get("__source_row__", "")
        warnings.append(_warning(source_file, source_row, "row", "Ignored empty row", ""))

    return df.loc[keep_indexes].copy(), warnings


def _find_column_candidates(columns):
    candidates = {field: [] for field in MAPPABLE_COLUMNS}

    for field, aliases in COLUMN_ALIASES.items():
        normalized_aliases = [_normalize_header(alias) for alias in aliases]
        for alias in normalized_aliases:
            for col in columns:
                if col == "__source_row__" or col in candidates[field]:
                    continue
                if col == alias:
                    candidates[field].append(col)

    for col in columns:
        if col == "__source_row__":
            continue
        if re.match(r"^manufacturer_\d+$", col) and col not in candidates["manufacturer"]:
            candidates["manufacturer"].append(col)
        if re.match(r"^manufacturer_part_number_\d+$", col) and col not in candidates["mpn"]:
            candidates["mpn"].append(col)
        if re.match(r"^supplier_\d+$", col) and col not in candidates["supplier"]:
            candidates["supplier"].append(col)
        if re.match(r"^supplier_part_number_\d+$", col) and col not in candidates["supplier_part_number"]:
            candidates["supplier_part_number"].append(col)
        if re.match(r"^supplier_unit_price_\d+$", col) and col not in candidates["unit_price"]:
            candidates["unit_price"].append(col)
        if re.match(r"^supplier_subtotal_\d+$", col) and col not in candidates["subtotal"]:
            candidates["subtotal"].append(col)
        if re.match(r"^manufacturer_lifecycle_\d+$", col) and col not in candidates["lifecycle_status"]:
            candidates["lifecycle_status"].append(col)

    return candidates


def _build_standard_rows(df, candidates, source_file):
    rows = []
    warnings = []
    multiplier_columns = _find_multiplier_columns(df.columns)

    for _, row in df.iterrows():
        notes = []
        source_row = row.get("__source_row__", "")

        output_row = {
            "source_file": source_file,
            "source_row": source_row,
        }

        selected_sources = {}
        for field in MAPPABLE_COLUMNS:
            value, source_column = _first_non_empty(row, candidates.get(field, []))
            output_row[field] = value
            selected_sources[field] = source_column

        output_row["build_multiplier"] = _format_multiplier_values(row, multiplier_columns)

        combined_notes = _apply_combined_manufacturer_mpn_parse(
            output_row,
            row,
            candidates,
            selected_sources,
        )
        notes.extend(combined_notes)

        if not output_row["mpn"]:
            name_value, name_column = _first_non_empty(row, ["name"])
            parsed_name = _parse_combined_manufacturer_mpn(name_value)

            if _mpn_is_clean(name_value) and selected_sources.get("mpn") != name_column:
                output_row["mpn"] = name_value
                notes.append("mpn filled from name fallback")
            elif parsed_name:
                parsed_manufacturer, parsed_mpn = parsed_name
                if not output_row["manufacturer"]:
                    output_row["manufacturer"] = parsed_manufacturer
                    notes.append(f"combined field parsed: manufacturer from {name_column}")
                output_row["mpn"] = parsed_mpn
                notes.append(f"combined field parsed: mpn from {name_column}")
            elif name_value:
                notes.append("manual review: name column may contain part number")

        output_row["notes"] = "; ".join(_dedupe_notes(notes))
        output_row["status"] = "manual_review"

        rows.append(output_row)

    clean_df = pd.DataFrame(rows, columns=STANDARD_COLUMNS)

    if multiplier_columns:
        warnings.append(
            _warning(
                source_file,
                "",
                "build_multiplier",
                "Detected build multiplier columns",
                ", ".join(multiplier_columns),
            )
        )

    return clean_df, warnings


def _clean_standard_rows(df, source_file):
    clean_df = df.copy().astype(object)
    warnings = []

    for col in STANDARD_COLUMNS:
        if col not in clean_df.columns:
            clean_df[col] = ""

    for col in clean_df.columns:
        clean_df[col] = clean_df[col].apply(clean_text)
    clean_df = clean_df.astype(object)

    for index, row in clean_df.iterrows():
        notes = _split_notes(row.get("notes", ""))
        source_row = row.get("source_row", "")

        declared_qty_text = clean_text(row.get("qty_per_board", ""))
        parsed_qty = _parse_quantity(declared_qty_text)
        designator_count = count_designators(row.get("designators", ""))

        if parsed_qty is not None:
            clean_df.at[index, "qty_per_board"] = parsed_qty
            if designator_count and int(float(parsed_qty)) != designator_count:
                notes.append(
                    f"qty mismatch: declared {parsed_qty} vs {designator_count} designators"
                )
                warnings.append(
                    _warning(
                        source_file,
                        source_row,
                        "qty_per_board",
                        "Declared quantity does not match designator count",
                        f"{parsed_qty} vs {designator_count}",
                    )
                )
        elif designator_count:
            clean_df.at[index, "qty_per_board"] = designator_count
            notes.append("quantity inferred from designators")
        elif declared_qty_text:
            notes.append(f"manual review: non-numeric quantity '{declared_qty_text}'")
            warnings.append(
                _warning(
                    source_file,
                    source_row,
                    "qty_per_board",
                    "Quantity is not numeric",
                    declared_qty_text,
                )
            )
        else:
            notes.append("manual review: missing quantity and designators")
            warnings.append(
                _warning(
                    source_file,
                    source_row,
                    "qty_per_board",
                    "Missing quantity and designators",
                    "",
                )
            )

        if not clean_text(row.get("mpn", "")):
            notes.append("missing mpn")
            warnings.append(_warning(source_file, source_row, "mpn", "Missing MPN", ""))

        if not clean_text(row.get("manufacturer", "")):
            notes.append("missing manufacturer")
            warnings.append(
                _warning(source_file, source_row, "manufacturer", "Missing manufacturer", "")
            )

        clean_df.at[index, "notes"] = "; ".join(_dedupe_notes(notes))
        clean_df.at[index, "status"] = determine_status(clean_df.loc[index])

    return clean_df[STANDARD_COLUMNS], warnings


def _apply_combined_manufacturer_mpn_parse(output_row, source_row, candidates, selected_sources):
    notes = []

    if output_row.get("manufacturer") and _mpn_is_clean(output_row.get("mpn", "")):
        return notes

    for source_column in candidates.get("mpn", []):
        raw_value = clean_text(source_row.get(source_column, ""))
        if not raw_value:
            continue

        parsed = _parse_combined_manufacturer_mpn(raw_value)
        if not parsed:
            continue

        parsed_manufacturer, parsed_mpn = parsed
        selected_mpn_source = selected_sources.get("mpn", "")

        if not output_row.get("manufacturer"):
            output_row["manufacturer"] = parsed_manufacturer
            notes.append(f"combined field parsed: manufacturer from {source_column}")

        if not output_row.get("mpn") or selected_mpn_source == source_column:
            output_row["mpn"] = parsed_mpn
            notes.append(f"combined field parsed: mpn from {source_column}")

        return notes

    return notes


def _parse_combined_manufacturer_mpn(value):
    text = clean_text(value)
    if not text or " " not in text:
        return None

    parts = text.split()
    if len(parts) < 2:
        return None

    mpn = parts[-1].strip(" ,;")
    manufacturer = " ".join(parts[:-1]).strip(" ,;")

    if not manufacturer or not _looks_like_part_number(mpn):
        return None

    if _is_generic_manufacturer_prefix(manufacturer):
        return None

    if _looks_like_part_number(manufacturer) and len(parts) == 2:
        return None

    return manufacturer, mpn


def _is_generic_manufacturer_prefix(value):
    normalized = clean_text(value).lower()
    return normalized in GENERIC_MANUFACTURER_PREFIXES


def _looks_like_part_number(value):
    text = clean_text(value)
    if len(text) < 2:
        return False
    if re.fullmatch(r"\d+\.\d+", text):
        return False
    if not re.search(r"\d", text):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._/+:-]*", text))


def _mpn_is_clean(value):
    text = clean_text(value)
    if not text:
        return False
    return " " not in text and _looks_like_part_number(text)


def _parse_quantity(value):
    text = clean_text(value)
    if not text:
        return None

    text = text.replace(",", "")
    match = re.fullmatch(r"(\d+)(?:\.0+)?", text)
    if match:
        return int(match.group(1))

    try:
        number = float(text)
    except ValueError:
        return None

    return int(number) if number.is_integer() else number


def _first_non_empty(row, columns):
    for col in columns:
        if col not in row.index:
            continue
        value = clean_text(row.get(col, ""))
        if value:
            return value, col
    return "", ""


def _find_multiplier_columns(columns):
    return [col for col in columns if MULTIPLIER_COLUMN_RE.match(col)]


def _format_multiplier_values(row, multiplier_columns):
    values = []
    for col in multiplier_columns:
        value = clean_text(row.get(col, ""))
        if value:
            values.append(f"{col}={value}")
    return "; ".join(values)


def _build_column_mapping_df(df, candidates, source_file):
    records = []

    for field in MAPPABLE_COLUMNS:
        candidate_columns = candidates.get(field, [])
        selected = candidate_columns[0] if candidate_columns else ""
        records.append(
            {
                "source_file": source_file,
                "standard_field": field,
                "selected_column": _original_column_name(df, selected),
                "candidate_columns": ", ".join(
                    _original_column_name(df, col) for col in candidate_columns
                ),
                "status": "mapped" if selected else "missing",
            }
        )

    multiplier_columns = _find_multiplier_columns(df.columns)
    if multiplier_columns:
        records.append(
            {
                "source_file": source_file,
                "standard_field": "build_multiplier",
                "selected_column": "",
                "candidate_columns": ", ".join(
                    _original_column_name(df, col) for col in multiplier_columns
                ),
                "status": "detected",
            }
        )

    return pd.DataFrame(
        records,
        columns=[
            "source_file",
            "standard_field",
            "selected_column",
            "candidate_columns",
            "status",
        ],
    )


def _build_mapped_columns(df, candidates):
    mapped = {}
    for field, candidate_columns in candidates.items():
        if candidate_columns:
            mapped[field] = _original_column_name(df, candidate_columns[0])
    return mapped


def _original_column_name(df, column):
    if not column:
        return ""
    return df.attrs.get("original_columns", {}).get(column, column)


def _make_warnings_df(warnings):
    if not warnings:
        return pd.DataFrame(columns=WARNING_COLUMNS)
    return pd.DataFrame(warnings, columns=WARNING_COLUMNS)


def _warning(source_file, source_row, field, message, value):
    return {
        "source_file": source_file,
        "source_row": source_row,
        "field": field,
        "message": message,
        "value": clean_text(value),
    }


def _split_notes(notes):
    return [note.strip() for note in clean_text(notes).split(";") if note.strip()]


def _dedupe_notes(notes):
    seen = set()
    deduped = []
    for note in notes:
        normalized = note.lower()
        if normalized in seen:
            continue
        seen.add(normalized)
        deduped.append(note)
    return deduped


def _format_workbook(output_path):
    workbook = load_workbook(output_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row > 1 and worksheet.max_column > 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            header = column_cells[0].column_letter
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[header].width = min(max(max_length + 2, 12), 60)

    workbook.save(output_path)
